import imaplib
import email
from email.header import decode_header
import pandas as pd
import os
import datetime
import re 
import uuid
from bs4 import BeautifulSoup
import tkinter as tk
from tkinter import ttk, messagebox
import configparser
import sys
from dateutil import parser as date_parser


def get_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))
    
def load_config():
    config = configparser.ConfigParser()
    config_path = os.path.join(get_base_dir(), 'email_config.ini')

    if os.path.exists(config_path):
        config.read(config_path)
    else:
        config['Credentials'] = {
            'email': '',
            'password': ''
        }
        config['Search'] = {
            'subject_keyword': 'Bukti Pembayaran Transaksi PT. KAI Persero',
            'unread_only': 'True'
        }
        config['Output'] = {
            'excel_file': 'email_attachment_report.xlsx',
            'attachments_dir': os.path.join(get_base_dir(), 'email_attachments')
        }
        with open(config_path, 'w') as f:
            config.write(f)
    return config, config_path

def save_config(config, config_path):
    with open(config_path, 'w') as f:
        config.write(f)

def create_attachments_dir(base_dir):
    attachments_dir = os.path.join(base_dir, 'email_attachments')
    os.makedirs(attachments_dir, exist_ok=True)
    return attachments_dir

def get_month_folder(attachments_dir, email_date):
    try:
        parsed_date = None
        if isinstance(email_date, str):
            try:
                # Use dateutil parser for more flexible date parsing
                parsed_date = date_parser.parse(email_date.split('(')[0].strip())
            except Exception:
                print(f"Couldn't parse date:{email_date}, using current date")
                parsed_date = datetime.datetime.now()
        else:
            # If email_date is already a datetime object
            parsed_date = email_date
            
        month_folder_name = parsed_date.strftime("%B %Y")
        month_folder_path = os.path.join(attachments_dir, month_folder_name)
        os.makedirs(month_folder_path, exist_ok=True)
        return month_folder_path
    except Exception as e:
        print(f"Error creating month folder: {e}")
        return attachments_dir

def save_attachment(part, attachments_dir, email_date=None):
    saved_attachments = []
    filename = part.get_filename()
    if not filename: 
        return saved_attachments
    try: 
        filename = decode_header(filename)[0][0]
        if isinstance(filename, bytes):
            filename = filename.decode('utf-8', errors='ignore')
    except Exception:
        filename = f"attachment_{uuid.uuid4()}"
    filename = re.sub(r'[^\w\-_\.]','_', filename)
    try:
        if email_date:
            save_dir = get_month_folder(attachments_dir, email_date)
        else:
            save_dir = attachments_dir
        
        filepath = os.path.join(save_dir, filename)
        counter = 1
        base, ext = os.path.splitext(filepath)
        while os.path.exists(filepath):
            filepath = f"{base}_{counter}{ext}"
            counter += 1
        
        with open(filepath, 'wb') as f:
            f.write(part.get_payload(decode=True))

        saved_attachments.append(filepath)
        return saved_attachments
    except Exception as e:
        print(f"Error saving attachment {filename}: {e}")
        return saved_attachments

def append_to_excel(new_data, output_file):
    try:
        if len(new_data) == 0:
            print("No email data to write to Excel")
            return "No new emails found to add to Excel"
            
        if not output_file.lower().endswith('.xlsx'):
            output_file = os.path.splitext(output_file)[0] + '.xlsx'
            
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
            
        print(f"Writing {len(new_data)} records to {output_file}")
        
        if os.path.exists(output_file):
            try:
                existing_df = pd.read_excel(output_file, sheet_name="Email Data")
                print(f"Loaded existing Excel with {len(existing_df)} records")
                
                for col in new_data.columns:
                    if col not in existing_df.columns:
                        existing_df[col] = ''
                
                combined_df = pd.concat([existing_df, new_data], ignore_index=True)
                print(f"Combined dataframe has {len(combined_df)} records")
                
                old_len = len(combined_df)
                combined_df.drop_duplicates(subset=['Subject', 'Sender', 'Date'], keep='first', inplace=True)
                print(f"After removing duplicates: {len(combined_df)} records (removed {old_len - len(combined_df)})")
                
                # Write data to the existing template
                combined_df.to_excel(output_file, sheet_name="Email Data", index=False)
                return f"Appended {len(new_data)} new emails to {output_file} (total: {len(combined_df)})"
            except Exception as e:
                print(f"Error reading existing Excel: {e}")
                print("Creating new file instead")
                # Create template first, then add data
                create_excel_template(output_file)
                new_data.to_excel(output_file, sheet_name="Email Data", index=False)
                return f"Created new file {output_file} with {len(new_data)} emails"
        else:
            print(f"Creating new Excel file at {output_file}")
            # Create template first, then add data
            create_excel_template(output_file)
            new_data.to_excel(output_file, sheet_name="Email Data", index=False)
            return f"Created new file {output_file} with {len(new_data)} emails"
    except Exception as e:
        print(f"Critical error in append_to_excel: {str(e)}")
        emergency_file = os.path.join(get_base_dir(), "emergency_email_data.xlsx")
        try:
            new_data.to_excel(emergency_file, index=False)
            return f"Error with original file. Data saved to {emergency_file}"
        except:
            return f"Critical error: Could not save email data anywhere. {str(e)}"
    
def create_excel_template(excel_file):
    """Create an Excel template for email data with proper formatting but no macros"""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Ensure the file is xlsx
        if excel_file.lower().endswith('.xlsm'):
            excel_file = excel_file.replace('.xlsm', '.xlsx')
            
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Email Data"
        
        # Define headers
        headers = ["Subject", "Sender", "To", "CC", "BCC", "Date", "Content", "Attachments", 
                  "Folder Path", "Tanggal Pengiriman", "Full Folder Path", "Folder Checklist", "SO Checklist"]
        
        # Add headers with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        column_widths = {
            1: 40,  # Subject
            2: 25,  # Sender
            3: 25,  # To
            4: 25,  # CC
            5: 20,  # BCC
            6: 20,  # Date
            7: 50,  # Content
            8: 30,  # Attachments
            9: 25,  # Folder Path
            10: 20, # Tanggal Pengiriman
            11: 30, # Full Folder Path
            12: 15, # Folder Checklist
            13: 15  # SO Checklist
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Save the template as xlsx
        wb.save(excel_file)
        
        print(f"Created template at {excel_file}")
        return True
    except Exception as e:
        print(f"Error creating template: {e}")
        return False

def clean_subject(subject):
    if subject: 
        decoded_subject = []
        for part, encoding in decode_header(subject):
            if isinstance(part, bytes):
                part = part.decode(encoding or 'utf-8', errors='ignore')
            decoded_subject.append(part)
        return ' '.join(decoded_subject)
    return ''

def extract_email_content(email_message):
    email_content = ""
    
    if isinstance(email_message, str):
        try:
            email_message = email.message_from_string(email_message)
        except Exception as e:
            print(f"Could not parse email message: {e}")
            return ""
    
    if email_message.is_multipart():
        for part in email_message.walk():
            content_type = part.get_content_type()
            
            if content_type == 'text/plain':
                try:
                    payload = part.get_payload(decode=True)
                    charset = part.get_content_charset('utf-8')
                    email_content += payload.decode(charset, errors='ignore') + "\n\n"
                except Exception as e:
                    print(f"Error decoding plain text part: {e}")
            
            elif content_type == 'text/html':
                try:
                    payload = part.get_payload(decode=True)
                    charset = part.get_content_charset('utf-8')
                    html_content = payload.decode(charset, errors='ignore')
                    
                    soup = BeautifulSoup(html_content, 'html.parser')
                    
                    for script in soup(["script", "style"]):
                        script.decompose()
                    
                    text = soup.get_text(separator=' ', strip=True)
                    
                    text = re.sub(r'\s+', ' ', text).strip()
                    
                    email_content += text + "\n\n"
                except Exception as e:
                    print(f"Error processing HTML content: {e}")
    
    else:
        content_type = email_message.get_content_type()
        try:
            payload = email_message.get_payload(decode=True)
            charset = email_message.get_content_charset('utf-8')
            
            if content_type == 'text/plain':
                email_content = payload.decode(charset, errors='ignore')
            elif content_type == 'text/html':
                soup = BeautifulSoup(payload.decode(charset, errors='ignore'), 'html.parser')
                
                for script in soup(["script", "style"]):
                    script.decompose()
                
                text = soup.get_text(separator=' ', strip=True)
                
                email_content = re.sub(r'\s+', ' ', text).strip()
        except Exception as e:
            print(f"Error processing single-part email: {e}")
    
    return email_content.strip()
    
def search_emails(
        mail,
        attachments_dir,
        subject_keyword=None,
        start_date=None, 
        end_date=None,
        unread_only=True,
        status_callback=None
):
    search_criteria = []
    if subject_keyword:
        search_criteria.append(f'SUBJECT "{subject_keyword}"')
        
    if start_date:
        search_criteria.append(f'SINCE "{start_date.strftime("%d-%b-%Y")}"')
    if end_date:
        search_criteria.append(f'BEFORE "{end_date.strftime("%d-%b-%Y")}"')
    if unread_only:
        search_criteria.append('UNSEEN')
    
    try:
        search_string = ' '.join(search_criteria) if search_criteria else 'ALL'
        if status_callback:
            status_callback(f"Executing IMAP search with criteria: {search_string}")
        
        result, data = mail.search(None, search_string)
        
        if not data[0]:
            if status_callback:
                status_callback("No emails found matching the search criteria")
            return []
            
        email_count = len(data[0].split())
        if status_callback:
            status_callback(f"Found {email_count} emails matching search criteria")
        
        email_list = []
        for i, num in enumerate(data[0].split()):
            if status_callback:
                status_callback(f"Processing email {i+1}/{email_count}")
            result, email_data = mail.fetch(num, '(RFC822)')
            try:
                raw_email = email_data[0][1]
                email_message = email.message_from_bytes(raw_email)
                email_subject = clean_subject(email_message['Subject'])
                email_sender = email_message['From']
                email_date = email_message['Date']
                email_to = email_message.get('To','')
                email_cc = email_message.get('CC', '')
                email_bcc = email_message.get('Bcc', '')
                email_content = extract_email_content(email_message)
                
                attachment_paths = []
                if email_message.is_multipart():
                    for part in email_message.walk():
                        if part.get_content_maintype() == 'multipart':
                            continue
                            
                        is_attachment = False
                        
                        if part.get('Content-Disposition') and part.get('Content-Disposition').startswith('attachment'):
                            is_attachment = True
                         
                        elif part.get_content_type() == 'application/octet-stream' or part.get_content_maintype() == 'application':
                            is_attachment = True
                            
                        elif part.get_filename():
                            is_attachment = True
                        
                        if is_attachment:
                            saved = save_attachment(part, attachments_dir, email_date)
                            if saved:
                                attachment_paths.extend(saved)
                                if status_callback:
                                    status_callback(f"  - Saved attachment: {os.path.basename(saved[0])}")
                
                email_list.append({
                    'Subject': email_subject,
                    'Sender': email_sender,
                    'To': email_to,
                    'CC': email_cc,
                    'BCC': email_bcc,
                    'Date': email_date, 
                    'Attachments': '; '.join(attachment_paths) if attachment_paths else '',
                    'Folder Path': '',
                    'Tanggal Pengiriman': '',
                    'Full Folder Path': '', 
                    'Folder Checklist': False,
                    'SO Checklist': False,
                    'Content': email_content,
                })
            except Exception as email_error:
                if status_callback:
                    status_callback(f"Error processing email {num}: {email_error}")
        
        if status_callback:
            status_callback(f"Successfully processed {len(email_list)} emails")
        return email_list

    except Exception as search_error:
        if status_callback:
            status_callback(f"Email search error: {search_error}")
        return []

class EmailProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title = 'Email Attachment Processor'
        self.root.geometry("700x600")
        self.config, self.config_path = load_config()
        
        self.tab_control = ttk.Notebook(root)
        self.setup_tab = ttk.Frame(self.tab_control)
        self.process_tab = ttk.Frame(self.tab_control)
        self.log_tab = ttk.Frame(self.tab_control)
        
        self.tab_control.add(self.setup_tab, text='Setup')
        self.tab_control.add(self.process_tab, text='Process Emails')
        self.tab_control.add(self.log_tab, text='Log')
        self.tab_control.pack(expand=1, fill="both")
        
        self.create_setup_tab()
        
        self.create_process_tab()
        
        self.create_log_tab()
        
        self.load_config_values()

    def create_folders_from_excel(self):
        try:
            excel_file = self.excel_var.get()
            if not excel_file:
                self.update_status("Please specify Excel file path first")
                return
        
            if excel_file.lower().endswith('.xlsm'):
                excel_file = excel_file.replace('.xlsm', '.xlsx')
                self.excel_var.set(excel_file)
                
            import pandas as pd
            from openpyxl import load_workbook
            import shutil
            
            df = pd.read_excel(excel_file)
            
            required_cols = ["Folder Path", "Tanggal Pengiriman", "Full Folder Path", "Folder Checklist", "Attachments"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                self.update_status(f"Missing columns in Excel: {', '.join(missing_cols)}")
                return
            
            # Get base path from settings
            base_path = self.base_path_var.get().strip()
            self.update_status(f"Using base path: '{base_path}'")
            if not base_path:
                self.update_status("Warning: Base Path for PO Folders is empty")
            
            folder_counter = 0
            attachment_counter = 0
            updated_rows = []
            
            for index, row in df.iterrows():
                attachments = row.get('Attachments', '')
                folder_checklist = row.get('Folder Checklist', False)
                
                # Skip rows without attachments or already processed
                if pd.isna(attachments) or not str(attachments).strip() or folder_checklist:
                    continue
                    
                folder_path = row.get('Folder Path', '')
                tanggal = row.get('Tanggal Pengiriman', '')
                
                # Convert values to strings, handling NaN/None values
                folder_path_str = str(folder_path).strip() if not pd.isna(folder_path) else ""
                tanggal_str = str(tanggal).strip() if not pd.isna(tanggal) else ""
                
                # Initialize path components list
                path_components = []
                
                # Add base path if it exists
                if base_path:
                    path_components.append(base_path)
                    
                # Add folder path if available
                if folder_path_str:
                    path_components.append(folder_path_str)
                    
                # Add tanggal if available
                if tanggal_str:
                    path_components.append(tanggal_str)
                    
                # Skip if we don't have enough components to create a meaningful path
                if len(path_components) < 2:  # Need at least basepath + one other component
                    self.update_status(f"Skipping row {index+2}: Insufficient path components")
                    continue
                    
                # Create full path
                full_path = os.path.join(*path_components)
                
                try:
                    # Create the directory
                    os.makedirs(full_path, exist_ok=True)
                    folder_counter += 1
                    self.update_status(f"Created folder: {full_path}")
                    
                    # Process attachments
                    moved_attachments = []
                    attachment_list = str(attachments).split('; ')
                    
                    for attachment_path in attachment_list:
                        attachment_path = attachment_path.strip()
                        if not attachment_path:
                            continue
                            
                        if os.path.exists(attachment_path):
                            try:
                                filename = os.path.basename(attachment_path)
                                dest_path = os.path.join(full_path, filename)
                                
                                # Move the file
                                shutil.move(attachment_path, dest_path)
                                
                                moved_attachments.append(dest_path)
                                attachment_counter += 1
                                self.update_status(f"  - Moved attachment: {filename} to {full_path}")
                            except Exception as e:
                                self.update_status(f"  - Error moving attachment {filename}: {e}")
                        else:
                            self.update_status(f"  - Attachment not found: {attachment_path}")
                    
                    # Add to update list if we moved any attachments
                    if moved_attachments:
                        updated_rows.append({
                            'index': index,
                            'full_path': full_path,
                            'moved_attachments': '; '.join(moved_attachments)
                        })
                    
                except Exception as e:
                    self.update_status(f"Error creating folder {full_path}: {e}")
            
            # Update Excel file with new information
            if updated_rows:
                wb = load_workbook(excel_file)
                ws = wb.active
            
                for row_data in updated_rows:
                    row_idx = row_data['index']
                    full_path = row_data['full_path']
                    moved_attachments = row_data['moved_attachments']
            
                    excel_row = row_idx + 2  # Adjusting for 1-based Excel rows and header
            
                    # Update Full Folder Path column (10)
                    ws.cell(row=excel_row, column=10).value = full_path
            
                    # Update Folder Checklist column (11) to mark as processed
                    ws.cell(row=excel_row, column=11).value = True
            
                    # Update Attachments column (7) with new paths
                    if moved_attachments:
                        ws.cell(row=excel_row, column=7).value = moved_attachments
                    
                wb.save(excel_file)
                
            self.update_status(f"Created {folder_counter} folders and moved {attachment_counter} attachments successfully")
        except Exception as e:
            self.update_status(f"Error processing folders: {e}")

    def update_server_settings(self, event=None):
        provider = self.email_provider_var.get()

        if provider == "Gmail":
            self.server_address_var.set('imap.gmail.com')
            self.server_port_var.set('993')
        elif provider =='Outlook/office365':
            self.server_address_var.set('outlook.office365.com')
            self.server_port_var.set('993')
        elif provider == 'Yahoo':
            self.server_address_var.set('imap.mail.yahoo.com')
            self.server_port_var.set('993')

    def create_setup_tab(self):
        cred_frame = ttk.LabelFrame(self.setup_tab, text="Email Credentials")
        cred_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(cred_frame, text="Email:").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
        self.email_var = tk.StringVar()
        ttk.Entry(cred_frame, textvariable=self.email_var, width=40).grid(column=1, row=0, padx=5, pady=5)
        
        ttk.Label(cred_frame, text="Password:").grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
        self.password_var = tk.StringVar()
        ttk.Entry(cred_frame, textvariable=self.password_var, show="*", width=40).grid(column=1, row=1, padx=5, pady=5)
        
        server_frame = ttk.LabelFrame(self.setup_tab, text="Mail Server Settings")
        server_frame.pack(fill="x", padx=10, pady=10)

        ttk.Label(server_frame, text="Email Provider:").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
        self.email_provider_var = tk.StringVar()
        provider_combo = ttk.Combobox(server_frame, textvariable=self.email_provider_var, width=20)
        provider_combo['values'] = ('Gmail', 'Outlook/Office365', 'Yahoo', 'Custom')
        provider_combo.grid(column=1, row=0, padx=5, pady=5)
        provider_combo.bind('<<ComboboxSelected>>', self.update_server_settings)

        ttk.Label(server_frame, text="IMAP Server:").grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
        self.server_address_var = tk.StringVar()
        ttk.Entry(server_frame, textvariable=self.server_address_var, width=40).grid(column=1, row=1, padx=5, pady=5)
        
        ttk.Label(server_frame, text="IMAP Port:").grid(column=0, row=2, sticky=tk.W, padx=5, pady=5)
        self.server_port_var = tk.StringVar(value="993")  # Default IMAP SSL port
        ttk.Entry(server_frame, textvariable=self.server_port_var, width=10).grid(column=1, row=2, sticky=tk.W, padx=5, pady=5)


        search_frame = ttk.LabelFrame(self.setup_tab, text="Search Criteria")
        search_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(search_frame, text="Subject Keyword (optional):").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
        self.subject_var = tk.StringVar()
        ttk.Entry(search_frame, textvariable=self.subject_var, width=40).grid(column=1, row=0, padx=5, pady=5)
        
        self.unread_var = tk.BooleanVar(value=True)  # Default to True
        ttk.Checkbutton(search_frame, text="Unread Only", variable=self.unread_var).grid(column=0, row=1, columnspan=2, sticky=tk.W, padx=5, pady=5)
        
        output_frame = ttk.LabelFrame(self.setup_tab, text="Output Settings")
        output_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(output_frame, text="Excel File:").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
        self.excel_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.excel_var, width=40).grid(column=1, row=0, padx=5, pady=5)
        
        ttk.Label(output_frame, text="Attachments Directory:").grid(column=0, row=1, sticky=tk.W, padx=5, pady=5)
        self.dir_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.dir_var, width=40).grid(column=1, row=1, padx=5, pady=5)
        
        ttk.Label(output_frame, text="Base Path for PO Folders:").grid(column=0, row=2, sticky=tk.W, padx=5, pady=5)
        self.base_path_var = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.base_path_var, width=40).grid(column=1, row=2, padx=5, pady=5)
        
        ttk.Button(self.setup_tab, text="Save Configuration", command=self.save_config_values).pack(pady=10)

        folder_button = ttk.Button(output_frame, text="Create Folders from Excel Data", 
                                command=self.create_folders_from_excel)
        folder_button.grid(column=0, row=4, columnspan=2, padx=5, pady=10)

    def create_process_tab(self):
        date_frame = ttk.LabelFrame(self.process_tab, text="Date Range")
        date_frame.pack(fill="x", padx=10, pady=10)
        
        ttk.Label(date_frame, text="Start Date (YYYY-MM-DD):").grid(column=0, row=0, sticky=tk.W, padx=5, pady=5)
        self.start_date_var = tk.StringVar()
        ttk.Entry(date_frame, textvariable=self.start_date_var, width=15).grid(column=1, row=0, padx=5, pady=5)
        
        ttk.Label(date_frame, text="End Date (YYYY-MM-DD):").grid(column=2, row=0, sticky=tk.W, padx=5, pady=5)
        self.end_date_var = tk.StringVar()
        ttk.Entry(date_frame, textvariable=self.end_date_var, width=15).grid(column=3, row=0, padx=5, pady=5)

        def set_today():
            today = datetime.datetime.now().strftime("%Y-%m-%d")
            self.end_date_var.set(today)
        
        ttk.Button(date_frame, text="Set Today", command=set_today).grid(column=4, row=0, padx=5, pady=5)

        ttk.Button(self.process_tab, text="Process Emails", command=self.process_emails).pack(pady=10)

        progress_frame = ttk.LabelFrame(self.process_tab, text="Progress")
        progress_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=5, pady=5)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        ttk.Label(progress_frame, textvariable=self.status_var).pack(padx=5, pady=5)
    
    def create_log_tab(self):
        log_frame = ttk.Frame(self.log_tab)
        log_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.log_text = tk.Text(log_frame, wrap=tk.WORD, height=25)
        self.log_text.pack(side=tk.LEFT, fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        ttk.Button(self.log_tab, text="Clear Log", command=self.clear_log).pack(pady=5)
    
    def load_config_values(self):
        if 'Credentials' in self.config:
            self.email_var.set(self.config['Credentials'].get('email', ''))
            self.password_var.set(self.config['Credentials'].get('password', ''))
        
        if 'Server' in self.config:
            self.email_provider_var.set(self.config['Server'].get('provider', 'Gmail'))
            self.server_address_var.set(self.config['Server'].get('address', 'imap.gmail.com'))
            self.server_port_var.set(self.config['Server'].get('port', '993'))

        if 'Search' in self.config:
            self.subject_var.set(self.config['Search'].get('subject_keyword', ''))
            self.unread_var.set(self.config['Search'].getboolean('unread_only', True))
        
        if 'Output' in self.config:
            self.excel_var.set(self.config['Output'].get('excel_file', 'email_attachment_report.xlsx'))
            self.dir_var.set(self.config['Output'].get('attachments_dir', ''))
            self.base_path_var.set(self.config['Output'].get('base_path', ''))
    
    def save_config_values(self):
        if 'Credentials' not in self.config:
            self.config['Credentials'] = {}
        self.config['Credentials']['email'] = self.email_var.get()
        self.config['Credentials']['password'] = self.password_var.get()

        if 'Server' not in self.config:
            self.config['Server'] = {}
        self.config['Server']['provider'] = self.email_provider_var.get()
        self.config['Server']['address'] = self.server_address_var.get()
        self.config['Server']['port'] = self.server_port_var.get()
        
        if 'Search' not in self.config:
            self.config['Search'] = {}
        self.config['Search']['subject_keyword'] = self.subject_var.get()
        self.config['Search']['unread_only'] = str(self.unread_var.get())
        
        if 'Output' not in self.config:
            self.config['Output'] = {}
        self.config['Output']['excel_file'] = self.excel_var.get()
        self.config['Output']['attachments_dir'] = self.dir_var.get()
        self.config['Output']['base_path'] = self.base_path_var.get()
        
        save_config(self.config, self.config_path)
        messagebox.showinfo("Configuration", "Configuration saved successfully!")
    
    def update_status(self, message):
        self.status_var.set(message)
        self.log_text.insert(tk.END, f"{datetime.datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def clear_log(self):
        self.log_text.delete(1.0, tk.END)

    def process_emails(self):
        try:
            start_date = None
            end_date = None
            
            if self.start_date_var.get().strip():
                try:
                    start_date = datetime.datetime.strptime(self.start_date_var.get().strip(), "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Date Error", "Invalid start date format. Use YYYY-MM-DD")
                    return
            
            if self.end_date_var.get().strip():
                try:
                    end_date = datetime.datetime.strptime(self.end_date_var.get().strip(), "%Y-%m-%d")
                    end_date += datetime.timedelta(days=1)
                except ValueError:
                    messagebox.showerror("Date Error", "Invalid end date format. Use YYYY-MM-DD")
                    return
            
            if not self.email_var.get() or not self.password_var.get():
                messagebox.showerror("Input Error", "Email and password are required")
                return
            
            email_provider = self.email_provider_var.get()
            server_address = self.server_address_var.get()
            server_port = self.server_port_var.get()

            if not server_address:
                server_address = 'imap.gmail.com'
                server_port = 993

            attachments_dir = self.dir_var.get()
            if not attachments_dir:
                attachments_dir = create_attachments_dir(get_base_dir())
            else:
                os.makedirs(attachments_dir, exist_ok=True)
            
            self.update_status("Connecting to email server : {server_address}")
            
            self.tab_control.select(self.log_tab)
            
            try:
                with imaplib.IMAP4_SSL(server_address, int(server_port)) as mail:
                    self.update_status("Logging in...")
                    mail.login(self.email_var.get(), self.password_var.get())
                    self.update_status("Connected successfully")
                    
                    mail.select('inbox')
                    self.update_status("Searching for emails...")
                    
                    emails = search_emails(
                        mail, 
                        attachments_dir,
                        subject_keyword=self.subject_var.get(),
                        start_date=start_date,
                        end_date=end_date, 
                        unread_only=self.unread_var.get(),
                        status_callback=self.update_status
                    )
                    
                    if emails:
                        df = pd.DataFrame(emails)
                        output_file = self.excel_var.get()
                        if not output_file:
                            output_file = os.path.join(get_base_dir(), 'email_attachment_report.xlsx')
                        
                        result = append_to_excel(df, output_file)
                        self.update_status(result)
                        messagebox.showinfo("Process Complete", f"Successfully processed {len(emails)} emails")
                    else:
                        self.update_status("No emails were found or processed")
                        messagebox.showinfo("Process Complete", "No emails were found matching your criteria")
            
            except imaplib.IMAP4.error as login_error:
                error_msg = f"IMAP Login Error: {login_error}"
                self.update_status(error_msg)
                messagebox.showerror("Login Error", error_msg)
            
            except Exception as e:
                error_msg = f"Unexpected error: {e}"
                self.update_status(error_msg)
                messagebox.showerror("Error", error_msg)
        
        except Exception as e:
            messagebox.showerror("Error", f"An unexpected error occurred: {e}")

def main():
    root = tk.Tk()
    app = EmailProcessorApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()